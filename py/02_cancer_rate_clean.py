from openpyxl import load_workbook
import pandas as pd
import numpy as np

for i in range(1, 1009):

    path1 = '/Users/nathanoliver/Desktop/Cancer Rates/xlsx/02_xlsx/incd-' + str(i) + '.xlsx'
    path2 = '/Users/nathanoliver/Desktop/Cancer Rates/xlsx/03_xlsx_modified/incd-' + \
        str(i) + '-modified.xlsx'

    print('loading workbook')
    wb = load_workbook(path1)
    print('loaded workbook')

    sheet = wb["Sheet"]

    sheet.cell(row=1, column=1).value = 'county'
    sheet.cell(row=1, column=2).value = 'state'
    sheet.cell(row=1, column=3).value = 'cancer'
    sheet.cell(row=1, column=4).value = 'race'
    sheet.cell(row=1, column=5).value = 'sex'
    sheet.cell(row=1, column=6).value = 'age'
    sheet.cell(row=1, column=7).value = 'incidence_rate'
    sheet.cell(row=1, column=8).value = 'lower_95%'
    sheet.cell(row=1, column=9).value = 'upper_95%'
    sheet.cell(row=1, column=10).value = 'average_annual_count'

    exit_text = 'Created by'

    txt = 'not cell'

    row = 0

    while exit_text not in txt:
        row = row + 1

        txt = sheet.cell(row=row, column=1).value

        try:
            cell = txt.split(',')
            for col in range(1, len(cell) + 1):
                sheet.cell(row=row, column=col).value = cell[col - 1]

                if col == 2:
                    state_org = cell[col - 1]
                    state_strip = state_org.strip()
                    state_split = state_strip.split('(')
                    state = state_split[0]
                    # print(state)
                    sheet.cell(row=row, column=col).value = state

        except AttributeError or TypeError:
            txt = 'not cell'

    wb.save(path2)
