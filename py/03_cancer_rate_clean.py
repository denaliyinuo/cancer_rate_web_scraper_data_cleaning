from openpyxl import load_workbook
import pandas as pd
import numpy as np

for i in range(951, 1009):
    print(i)
    path2 = '/Users/nathanoliver/Desktop/Cancer Rates/xlsx/03_xlsx_modified/incd-' + \
        str(i) + '-modified.xlsx'
    path3 = '/Users/nathanoliver/Desktop/Cancer Rates/xlsx/04_xlsx_final/incd-' + \
        str(i) + '-final.xlsx'

    print('loading workbook')
    wb = load_workbook(path2)
    print('loaded workbook')

    sheet_old = wb["Sheet"]
    sheet = wb.create_sheet("Sheet1", 1)

    sheet.cell(row=1, column=1).value = 'county'
    sheet.cell(row=1, column=2).value = 'state'
    sheet.cell(row=1, column=3).value = 'cancer'
    sheet.cell(row=1, column=4).value = 'race'
    sheet.cell(row=1, column=5).value = 'sex'
    sheet.cell(row=1, column=6).value = 'age'
    sheet.cell(row=1, column=7).value = 'incidence_rate'
    sheet.cell(row=1, column=8).value = 'lower_95%'
    sheet.cell(row=1, column=9).value = 'upper_95%'
    sheet.cell(row=1, column=10).value = 'average_annual_count'

    cancer = sheet_old.cell(row=3, column=1).value.strip()
    race = sheet_old.cell(row=5, column=1).value.strip()
    sex = sheet_old.cell(row=5, column=2).value.strip()
    if cancer == 'Childhood (Ages <15':
        age = '<15'
        cancer = 'Childhood - Ages <15'
    elif cancer == 'Childhood (Ages <20':
        age = '<20'
        cancer = 'Childhood - Ages <20'
    else:
        age = sheet_old.cell(row=5, column=3).value.strip()

    for row in range(10, 3152):

        test_cell = sheet_old.cell(row=row, column=3).value
        test_cell_backup = sheet_old.cell(row=row, column=2).value

        # cancer
        sheet.cell(row=row - 8, column=3).value = cancer
        # race
        sheet.cell(row=row - 8, column=4).value = race
        # sex
        sheet.cell(row=row - 8, column=5).value = sex
        # age
        sheet.cell(row=row - 8, column=6).value = age

        # county
        if test_cell_backup == '11001' or test_cell_backup == '72001':
            if test_cell_backup == '11001':
                # print('Entered DC')
                sheet.cell(
                    row=row - 8, column=1).value = 'District of Columbia'
                sheet.cell(
                    row=row - 8, column=2).value = 'District of Columbia'
            elif test_cell_backup == '72001':
                sheet.cell(row=row - 8, column=1).value = 'Puerto Rico'
                sheet.cell(row=row - 8, column=2).value = 'Puerto Rico'

            # incidence_rate
            sheet.cell(
                row=row - 8, column=7).value = sheet_old.cell(row=row, column=4).value.replace('#', '').strip()
            # lower_95
            sheet.cell(
                row=row - 8, column=8).value = sheet_old.cell(row=row, column=5).value.strip()
            # upper_95
            sheet.cell(
                row=row - 8, column=9).value = sheet_old.cell(row=row, column=6).value.strip()
            # average_annual_count
            sheet.cell(
                row=row - 8, column=10).value = sheet_old.cell(row=row, column=7).value.strip()
        else:

            sheet.cell(
                row=row - 8, column=1).value = sheet_old.cell(row=row, column=1).value.strip()
            # print(sheet.cell(row=row - 8, column=1).value)
            # state
            sheet.cell(
                row=row - 8, column=2).value = sheet_old.cell(row=row, column=2).value.strip()

            # print(test_cell)

        try:
            # print(len(test_cell) == 5)
            # print(test_cell.isnumeric() == True)

            if (len(test_cell) == 5) and (test_cell.isnumeric() == True):
                # print('Entered True')
                # print()
                # incidence_rate
                sheet.cell(
                    row=row - 8, column=7).value = sheet_old.cell(row=row, column=5).value.replace('#', '').strip()
                # lower_95
                sheet.cell(
                    row=row - 8, column=8).value = sheet_old.cell(row=row, column=6).value.strip()
                # upper_95
                sheet.cell(
                    row=row - 8, column=9).value = sheet_old.cell(row=row, column=7).value.strip()
                # average_annual_count
                sheet.cell(
                    row=row - 8, column=10).value = sheet_old.cell(row=row, column=8).value.strip()
            elif (test_cell_backup != '11001'):
                if test_cell_backup != '72001':
                    # print('Entered False1')
                    # print(test_cell_backup)
                    # print()
                    # incidence_rate
                    sheet.cell(
                        row=row - 8, column=7).value = sheet_old.cell(row=row, column=6).value.replace('#', '').strip()
                    # lower_95
                    sheet.cell(
                        row=row - 8, column=8).value = sheet_old.cell(row=row, column=7).value.strip()
                    # upper_95
                    sheet.cell(
                        row=row - 8, column=9).value = sheet_old.cell(row=row, column=8).value.strip()
                    # average_annual_count
                    sheet.cell(
                        row=row - 8, column=10).value = sheet_old.cell(row=row, column=9).value.strip()
        except TypeError:
            # print('Entered False2')
            # print()
            # incidence_rate
            sheet.cell(
                row=row - 8, column=7).value = sheet_old.cell(row=row, column=6).value.strip()
            # lower_95
            sheet.cell(
                row=row - 8, column=8).value = sheet_old.cell(row=row, column=7).value.strip()
            # upper_95
            sheet.cell(
                row=row - 8, column=9).value = sheet_old.cell(row=row, column=8).value.strip()
            # average_annual_count
            sheet.cell(
                row=row - 8, column=10).value = sheet_old.cell(row=row, column=9).value.strip()
        # set columns to blank cells
        for delete in range(11, 14):
            sheet.cell(row=row - 8, column=delete).value = ''

    sheet.cell(row=2, column=1).value = 'US'
    sheet.cell(row=2, column=2).value = 'US'

    sheet.cell(row=row - 8, column=10).value

    wb.remove(sheet_old)

    wb.save(path3)
